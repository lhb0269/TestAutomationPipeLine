import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 테스트 케이스 데이터 생성
test_cases = [
    # 정상 테스트 케이스
    {
        'TC ID': 'No.1',
        'Category': '정상 테스트',
        'Title': '유효한 넥슨 계정으로 로그인 성공',
        'Precondition': 'FC온라인 메인 페이지 접속, 유효한 넥슨 계정 보유 ( / )',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: \n3. 비밀번호 입력: \n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '로그인 성공 후 https://fconline.nexon.com/main/index#today 페이지로 이동',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.2',
        'Category': '정상 테스트',
        'Title': '로그인 상태 유지 체크박스 선택 후 로그인',
        'Precondition': 'FC온라인 메인 페이지 접속, 유효한 넥슨 계정 보유',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: \n3. 비밀번호 입력: \n4. 로그인 상태 유지 체크박스 선택\n5. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '로그인 성공 및 로그인 상태 유지 설정',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.3',
        'Category': '정상 테스트',
        'Title': 'Naver SNS 계정으로 로그인 성공',
        'Precondition': 'FC온라인 메인 페이지 접속, 유효한 네이버 계정 보유 ( / )',
        'Test Step': '1. 로그인 모달 호출\n2. Naver 계정으로 로그인 버튼 클릭\n3. 네이버 로그인 페이지에서 계정 정보 입력\n4. 네이버 로그인 완료',
        'Expected Result': 'FC온라인으로 돌아와 로그인 상태 확인',
        'Priority': 'Medium'
    },
    # 비정상 테스트 케이스
    {
        'TC ID': 'No.4',
        'Category': '비정상 테스트',
        'Title': '아이디 필드 미입력 상태로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 필드 공백으로 유지\n3. 비밀번호 입력: \n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '아이디를 입력해주세요 또는 유사한 에러 메시지 표시',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.5',
        'Category': '비정상 테스트',
        'Title': '비밀번호 필드 미입력 상태로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: \n3. 비밀번호 필드 공백으로 유지\n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '비밀번호를 입력해주세요 또는 유사한 에러 메시지 표시',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.6',
        'Category': '비정상 테스트',
        'Title': '아이디/비밀번호 모두 미입력 상태로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 필드 공백으로 유지\n3. 비밀번호 필드 공백으로 유지\n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '아이디와 비밀번호를 입력해주세요 또는 유사한 에러 메시지 표시',
        'Priority': 'Medium'
    },
    {
        'TC ID': 'No.7',
        'Category': '비정상 테스트',
        'Title': '잘못된 아이디로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: invalid_id@test.com\n3. 비밀번호 입력: \n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '존재하지 않는 아이디입니다 또는 로그인 정보가 올바르지 않습니다 에러 메시지 표시',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.8',
        'Category': '비정상 테스트',
        'Title': '잘못된 비밀번호로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: \n3. 비밀번호 입력: wrong_password\n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '비밀번호가 올바르지 않습니다 또는 로그인 정보가 올바르지 않습니다 에러 메시지 표시',
        'Priority': 'High'
    },
    {
        'TC ID': 'No.9',
        'Category': '비정상 테스트',
        'Title': '아이디/비밀번호 모두 잘못된 정보로 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: invalid_user@test.com\n3. 비밀번호 입력: wrong_password\n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '로그인 정보가 올바르지 않습니다 에러 메시지 표시',
        'Priority': 'Medium'
    },
    {
        'TC ID': 'No.10',
        'Category': '비정상 테스트',
        'Title': 'SQL Injection 시도 (보안 테스트)',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': "1. 로그인 모달 호출\n2. 아이디 입력: admin' OR '1'='1\n3. 비밀번호 입력: admin' OR '1'='1\n4. 넥슨ID 로그인 버튼 클릭",
        'Expected Result': 'SQL Injection이 차단되고 적절한 에러 메시지 또는 로그인 실패 처리',
        'Priority': 'Low'
    },
    # 예외 테스트 케이스 (주요한 것들만)
    {
        'TC ID': 'No.11',
        'Category': '예외 테스트',
        'Title': '네트워크 연결 해제 상태에서 로그인 시도',
        'Precondition': 'FC온라인 메인 페이지 접속, 네트워크 연결 차단',
        'Test Step': '1. 로그인 모달 호출\n2. 아이디 입력: \n3. 비밀번호 입력: \n4. 넥슨ID 로그인 버튼 클릭',
        'Expected Result': '네트워크 연결을 확인해주세요 또는 연결 오류 메시지 표시',
        'Priority': 'Medium'
    },
    {
        'TC ID': 'No.12',
        'Category': '예외 테스트',
        'Title': '로그인 모달 로딩 중 페이지 새로고침',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. 모달 로딩 중 페이지 새로고침 (F5)',
        'Expected Result': '메인 페이지로 정상 복귀, 모달 닫힘 상태 유지',
        'Priority': 'Medium'
    },
    {
        'TC ID': 'No.13',
        'Category': '예외 테스트',
        'Title': 'SNS 로그인 중 인증 취소',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 모달 호출\n2. Naver 계정으로 로그인 버튼 클릭\n3. 네이버 로그인 페이지에서 취소 버튼 클릭',
        'Expected Result': 'FC온라인 로그인 모달로 정상 복귀',
        'Priority': 'Low'
    },
    {
        'TC ID': 'No.14',
        'Category': '예외 테스트',
        'Title': '로그인 시도 연속 실패 (5회 이상)',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1-5. 잘못된 정보로 로그인 시도 5회 반복\n6. 6번째 로그인 시도',
        'Expected Result': '계정 잠금 또는 CAPTCHA 인증 요구, 로그인 시도 횟수 초과 메시지',
        'Priority': 'Medium'
    },
    # 성능 테스트 케이스
    {
        'TC ID': 'No.15',
        'Category': '성능 테스트',
        'Title': '로그인 모달 로딩 시간 측정',
        'Precondition': 'FC온라인 메인 페이지 접속',
        'Test Step': '1. 로그인 버튼 클릭하여 모달 호출\n2. 모달 완전 로딩까지 시간 측정',
        'Expected Result': '3초 이내 로그인 모달 완전 로딩',
        'Priority': 'Low'
    },
    {
        'TC ID': 'No.16',
        'Category': '성능 테스트',
        'Title': '로그인 처리 응답 시간 측정',
        'Precondition': '로그인 모달 표시, 유효한 계정 정보 준비',
        'Test Step': '1. 아이디/비밀번호 입력\n2. 로그인 버튼 클릭\n3. 로그인 완료까지 시간 측정',
        'Expected Result': '5초 이내 로그인 처리 완료',
        'Priority': 'Low'
    }
]

# DataFrame 생성
df = pd.DataFrame(test_cases)

# Excel 파일로 저장
with pd.ExcelWriter('C:/AIChallenge(QA)/docs/fc-online-test-cases.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='FC온라인 테스트케이스', index=False)
    
    # 워크시트 가져오기
    worksheet = writer.sheets['FC온라인 테스트케이스']
    
    # 헤더 스타일링
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 카테고리별 색상 설정
    category_colors = {
        '정상 테스트': 'D5E8D4',
        '비정상 테스트': 'FFE6CC', 
        '예외 테스트': 'F8CECC',
        '성능 테스트': 'E1D5E7'
    }
    
    # 각 행에 카테고리별 색상 적용
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
        category = df.iloc[row_num-2]['Category']
        if category in category_colors:
            fill_color = PatternFill(start_color=category_colors[category], 
                                   end_color=category_colors[category], 
                                   fill_type='solid')
            for cell in row:
                cell.fill = fill_color
    
    # 열 너비 자동 조정
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 텍스트 줄 바꿈 설정
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

print('FC온라인 테스트 케이스 Excel 파일이 성공적으로 생성되었습니다.')
print('파일 위치: C:/AIChallenge(QA)/docs/fc-online-test-cases.xlsx')
