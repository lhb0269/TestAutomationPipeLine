import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_test_cases_excel():
    # 테스트 케이스 데이터 정의
    test_cases = [
        # 정상 시나리오
        {
            "TC ID": "No.1",
            "Category": "정상 시나리오",
            "Title": "유효한 계정 정보로 로그인이 성공하는지 확인",
            "Precondition": "- 브라우저가 실행된 상태\n- 인터넷 연결이 정상인 상태\n- 테스트 계정이 활성화된 상태",
            "Test Step": "1. https://m.albamon.com/user-account/login?my_page=1 페이지 접속\n2. 아이디 입력 필드에 '' 입력\n3. 비밀번호 입력 필드에 '' 입력\n4. 로그인 버튼 클릭",
            "Expected Result": "- 로그인이 성공하여 /personal/mypage 페이지로 이동\n- 마이페이지가 정상적으로 표시됨",
            "Priority": "High"
        },
        {
            "TC ID": "No.2",
            "Category": "정상 시나리오",
            "Title": "로그인 페이지에 정상적으로 접근할 수 있는지 확인",
            "Precondition": "- 브라우저가 실행된 상태\n- 인터넷 연결이 정상인 상태",
            "Test Step": "1. https://m.albamon.com/user-account/login?my_page=1 페이지 접속\n2. 페이지 로딩 완료 대기\n3. 로그인 폼 요소 확인",
            "Expected Result": "- 로그인 페이지가 정상적으로 로딩됨\n- 아이디 입력 필드가 표시됨\n- 비밀번호 입력 필드가 표시됨\n- 로그인 버튼이 표시됨\n- SNS 로그인 버튼들이 표시됨",
            "Priority": "High"
        },
        {
            "TC ID": "No.3",
            "Category": "정상 시나리오",
            "Title": "로그인 폼의 입력 필드가 정상적으로 작동하는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드 클릭\n2. 'testuser' 입력\n3. 비밀번호 입력 필드 클릭\n4. 'testpass' 입력\n5. 입력된 값 확인",
            "Expected Result": "- 아이디 필드에 입력된 텍스트가 정상 표시됨\n- 비밀번호 필드에 입력된 텍스트가 마스킹되어 표시됨\n- 입력 포커스가 정상적으로 이동됨",
            "Priority": "Medium"
        },
        
        # 비정상 시나리오
        {
            "TC ID": "No.4",
            "Category": "비정상 시나리오",
            "Title": "아이디를 입력하지 않고 로그인 시도 시 검증 메시지가 표시되는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드를 비운 상태로 유지\n2. 비밀번호 입력 필드에 '' 입력\n3. 로그인 버튼 클릭",
            "Expected Result": "- '아이디를 입력해주세요.' 또는 유사한 검증 메시지 표시 또는 다이얼로그 팝업 노출\n- 로그인이 진행되지 않음\n- 페이지 이동이 발생하지 않음",
            "Priority": "High"
        },
        {
            "TC ID": "No.5",
            "Category": "비정상 시나리오",
            "Title": "비밀번호를 입력하지 않고 로그인 시도 시 검증 메시지가 표시되는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드에 '' 입력\n2. 비밀번호 입력 필드를 비운 상태로 유지\n3. 로그인 버튼 클릭",
            "Expected Result": "- '비밀번호를 입력해주세요.' 또는 유사한 검증 메시지 표시 또는 다이얼로그 팝업 노출\n- 로그인이 진행되지 않음\n- 페이지 이동이 발생하지 않음",
            "Priority": "High"
        },
        {
            "TC ID": "No.6",
            "Category": "비정상 시나리오",
            "Title": "아이디와 비밀번호를 모두 입력하지 않고 로그인 시도 시 검증 메시지가 표시되는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드를 비운 상태로 유지\n2. 비밀번호 입력 필드를 비운 상태로 유지\n3. 로그인 버튼 클릭",
            "Expected Result": "- '아이디를 입력해주세요.' 또는 '아이디와 비밀번호를 입력해주세요.' 검증 메시지 표시 또는 다이얼로그 팝업 노출\n- 로그인이 진행되지 않음",
            "Priority": "High"
        },
        {
            "TC ID": "No.7",
            "Category": "비정상 시나리오",
            "Title": "존재하지 않는 아이디로 로그인 시도 시 에러 메시지가 표시되는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드에 'invaliduser123' 입력\n2. 비밀번호 입력 필드에 '' 입력\n3. 로그인 버튼 클릭\n4. 서버 응답 대기",
            "Expected Result": "- '아이디 또는 비밀번호가 일치하지 않습니다.' 또는 유사한 에러 메시지 표시 또는 다이얼로그 팝업 노출\n- 로그인이 실패함\n- 로그인 페이지에 머물러 있음",
            "Priority": "High"
        },
        {
            "TC ID": "No.8",
            "Category": "비정상 시나리오",
            "Title": "올바른 아이디와 잘못된 비밀번호로 로그인 시도 시 에러 메시지가 표시되는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드에 '' 입력\n2. 비밀번호 입력 필드에 'wrongpassword' 입력\n3. 로그인 버튼 클릭\n4. 서버 응답 대기",
            "Expected Result": "- '아이디 또는 비밀번호가 일치하지 않습니다.' 또는 유사한 에러 메시지 표시 또는 다이얼로그 팝업 노출\n- 로그인이 실패함\n- 로그인 페이지에 머물러 있음",
            "Priority": "High"
        },
        {
            "TC ID": "No.9",
            "Category": "비정상 시나리오",
            "Title": "특수문자나 스크립트가 포함된 입력에 대한 처리 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드에 '<script>alert(\"test\")</script>' 입력\n2. 비밀번호 입력 필드에 'SELECT * FROM users' 입력\n3. 로그인 버튼 클릭",
            "Expected Result": "- 특수문자가 적절히 이스케이프 처리됨\n- XSS나 SQL Injection 공격이 차단됨\n- 로그인 실패 메시지 표시",
            "Priority": "Medium"
        },
        
        # 예외 시나리오
        {
            "TC ID": "No.10",
            "Category": "예외 시나리오",
            "Title": "네트워크 연결이 끊어진 상태에서 로그인 시도 시 처리 확인",
            "Precondition": "- 로그인 페이지가 로딩된 상태\n- 네트워크 연결을 끊을 수 있는 환경",
            "Test Step": "1. 아이디 입력 필드에 '' 입력\n2. 비밀번호 입력 필드에 '' 입력\n3. 네트워크 연결 차단\n4. 로그인 버튼 클릭\n5. 타임아웃 시간 대기",
            "Expected Result": "- '네트워크 연결을 확인해주세요.' 또는 유사한 에러 메시지 표시 또는 다이얼로그 팝업 노출\n- 적절한 로딩 인디케이터 표시\n- 타임아웃 후 에러 처리",
            "Priority": "Medium"
        },
        {
            "TC ID": "No.11",
            "Category": "예외 시나리오",
            "Title": "서버 응답이 지연되는 상황에서의 사용자 경험 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태\n- 서버 응답 지연을 시뮬레이션할 수 있는 환경",
            "Test Step": "1. 아이디 입력 필드에 '' 입력\n2. 비밀번호 입력 필드에 '' 입력\n3. 로그인 버튼 클릭\n4. 30초 동안 대기하며 UI 변화 관찰",
            "Expected Result": "- 로딩 스피너나 인디케이터 표시\n- 로그인 버튼 비활성화\n- 적절한 시간 후 타임아웃 처리 또는 응답 처리",
            "Priority": "Medium"
        },
        {
            "TC ID": "No.12",
            "Category": "예외 시나리오",
            "Title": "브라우저 뒤로가기 버튼 사용 후 로그인 페이지 재접근 시 동작 확인",
            "Precondition": "- 로그인 페이지에서 다른 페이지로 이동한 상태",
            "Test Step": "1. 로그인 페이지 접속\n2. 알바몬 메인 페이지로 이동\n3. 브라우저 뒤로가기 버튼 클릭\n4. 로그인 폼 상태 확인",
            "Expected Result": "- 로그인 페이지가 정상적으로 표시됨\n- 입력 필드가 초기화된 상태\n- 모든 기능이 정상적으로 동작함",
            "Priority": "Low"
        },
        {
            "TC ID": "No.13",
            "Category": "예외 시나리오",
            "Title": "로그인 페이지에서 새로고침 후 기능 동작 확인",
            "Precondition": "- 로그인 페이지가 로딩된 상태\n- 일부 입력값이 입력된 상태",
            "Test Step": "1. 아이디 입력 필드에 'test' 입력\n2. F5 키 또는 새로고침 버튼 클릭\n3. 페이지 로딩 완료 대기\n4. 입력 필드 상태 확인",
            "Expected Result": "- 페이지가 정상적으로 새로고침됨\n- 모든 입력 필드가 초기화됨\n- 로그인 기능이 정상적으로 동작함",
            "Priority": "Low"
        },
        {
            "TC ID": "No.14",
            "Category": "예외 시나리오",
            "Title": "로그인 버튼을 빠르게 여러 번 클릭했을 때 처리 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 아이디 입력 필드에 '' 입력\n2. 비밀번호 입력 필드에 '' 입력\n3. 로그인 버튼을 빠르게 5번 연속 클릭\n4. 서버 응답 및 UI 변화 확인",
            "Expected Result": "- 첫 번째 클릭 후 버튼이 비활성화됨\n- 중복 요청이 방지됨\n- 한 번의 로그인 처리만 수행됨",
            "Priority": "Medium"
        },
        
        # SNS 로그인 시나리오
        {
            "TC ID": "No.15",
            "Category": "SNS 로그인",
            "Title": "카카오 로그인 버튼 클릭 시 카카오 인증 페이지로 이동하는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태\n- 카카오 로그인 버튼이 표시된 상태",
            "Test Step": "1. 카카오 로그인 버튼 위치 확인\n2. 카카오 로그인 버튼 클릭\n3. 페이지 이동 확인\n4. 카카오 인증 페이지 로딩 확인",
            "Expected Result": "- 카카오 인증 페이지로 정상 이동\n- 카카오 로그인 폼이 표시됨\n- 알바몬으로 돌아가기 옵션 제공",
            "Priority": "Medium"
        },
        {
            "TC ID": "No.16",
            "Category": "SNS 로그인",
            "Title": "네이버 로그인 버튼 클릭 시 네이버 인증 페이지로 이동하는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태\n- 네이버 로그인 버튼이 표시된 상태",
            "Test Step": "1. 네이버 로그인 버튼 위치 확인\n2. 네이버 로그인 버튼 클릭\n3. 페이지 이동 확인\n4. 네이버 인증 페이지 로딩 확인",
            "Expected Result": "- 네이버 인증 페이지로 정상 이동\n- 네이버 로그인 폼이 표시됨\n- 알바몬으로 돌아가기 옵션 제공",
            "Priority": "Medium"
        },
        {
            "TC ID": "No.17",
            "Category": "SNS 로그인",
            "Title": "카카오 로그인 인증 과정에서 취소 버튼 클릭 시 처리 확인",
            "Precondition": "- 카카오 인증 페이지가 로딩된 상태",
            "Test Step": "1. 카카오 로그인 버튼 클릭하여 인증 페이지 이동\n2. 카카오 인증 페이지에서 취소 또는 뒤로가기\n3. 알바몬 로그인 페이지로 돌아가는지 확인",
            "Expected Result": "- 알바몬 로그인 페이지로 정상 복귀\n- 에러 메시지 없이 정상 상태 유지\n- 다시 로그인 시도 가능한 상태",
            "Priority": "Low"
        },
        {
            "TC ID": "No.18",
            "Category": "SNS 로그인",
            "Title": "네이버 로그인 인증 과정에서 취소 버튼 클릭 시 처리 확인",
            "Precondition": "- 네이버 인증 페이지가 로딩된 상태",
            "Test Step": "1. 네이버 로그인 버튼 클릭하여 인증 페이지 이동\n2. 네이버 인증 페이지에서 취소 또는 뒤로가기\n3. 알바몬 로그인 페이지로 돌아가는지 확인",
            "Expected Result": "- 알바몬 로그인 페이지로 정상 복귀\n- 에러 메시지 없이 정상 상태 유지\n- 다시 로그인 시도 가능한 상태",
            "Priority": "Low"
        },
        {
            "TC ID": "No.19",
            "Category": "SNS 로그인",
            "Title": "카카오 계정으로 로그인이 성공하는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태\n- 카카오 테스트 계정 (@naver.com, ) 사용 가능",
            "Test Step": "1. 카카오 로그인 버튼 클릭\n2. 카카오 인증 페이지 이동\n3. 카카오 계정 정보 입력 (@naver.com, )\n4. 카카오 로그인 버튼 클릭\n5. 알바몬 마이페이지 이동 확인",
            "Expected Result": "- 카카오 로그인이 성공하여 /personal/mypage 페이지로 이동\n- 마이페이지가 정상적으로 표시됨\n- 카카오 계정으로 연동된 사용자 정보 표시",
            "Priority": "High"
        },
        {
            "TC ID": "No.20",
            "Category": "SNS 로그인",
            "Title": "네이버 계정으로 로그인이 성공하는지 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태\n- 네이버 테스트 계정 (, ) 사용 가능",
            "Test Step": "1. 네이버 로그인 버튼 클릭\n2. 네이버 인증 페이지 이동\n3. 네이버 계정 정보 입력 (, )\n4. 네이버 로그인 버튼 클릭\n5. 알바몬 마이페이지 이동 확인",
            "Expected Result": "- 네이버 로그인이 성공하여 /personal/mypage 페이지로 이동\n- 마이페이지가 정상적으로 표시됨\n- 네이버 계정으로 연동된 사용자 정보 표시",
            "Priority": "High"
        },
        {
            "TC ID": "No.21",
            "Category": "SNS 로그인",
            "Title": "잘못된 카카오 계정으로 로그인 시도 시 에러 처리 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 카카오 로그인 버튼 클릭\n2. 카카오 인증 페이지 이동\n3. 잘못된 카카오 계정 정보 입력 (invalid@test.com, wrongpass)\n4. 카카오 로그인 버튼 클릭\n5. 에러 메시지 확인",
            "Expected Result": "- 카카오 인증 페이지에서 로그인 실패 메시지 표시\n- 알바몬 로그인 페이지로 복귀하지 않음\n- 재시도 가능한 상태 유지",
            "Priority": "Medium"
        },
        {
            "TC ID": "No.22",
            "Category": "SNS 로그인",
            "Title": "잘못된 네이버 계정으로 로그인 시도 시 에러 처리 확인",
            "Precondition": "- 로그인 페이지가 정상적으로 로딩된 상태",
            "Test Step": "1. 네이버 로그인 버튼 클릭\n2. 네이버 인증 페이지 이동\n3. 잘못된 네이버 계정 정보 입력 (invaliduser, wrongpass)\n4. 네이버 로그인 버튼 클릭\n5. 에러 메시지 확인",
            "Expected Result": "- 네이버 인증 페이지에서 로그인 실패 메시지 표시\n- 알바몬 로그인 페이지로 복귀하지 않음\n- 재시도 가능한 상태 유지",
            "Priority": "Medium"
        }
    ]
    
    # DataFrame 생성
    df = pd.DataFrame(test_cases)
    
    # Excel 파일 생성
    excel_file = 'C:/AIChallenge(QA)/docs/albamon_login_test_cases_updated.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='테스트 케이스', index=False)
        
        # 워크시트 가져오기
        worksheet = writer.sheets['테스트 케이스']
        
        # 스타일 설정
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # 카테고리별 색상 설정
        category_colors = {
            '정상 시나리오': PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid'),
            '비정상 시나리오': PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid'),
            '예외 시나리오': PatternFill(start_color='F0E8FF', end_color='F0E8FF', fill_type='solid'),
            'SNS 로그인': PatternFill(start_color='E8F8FF', end_color='E8F8FF', fill_type='solid')
        }
        
        # 우선순위별 글자 색상
        priority_colors = {
            'High': Font(color='FF0000', bold=True),
            'Medium': Font(color='FF8C00', bold=True),
            'Low': Font(color='008000', bold=True)
        }
        
        # 헤더 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 데이터 행 스타일 적용
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            category = worksheet[f'B{row_num}'].value
            priority = worksheet[f'G{row_num}'].value
            
            # 카테고리별 배경색 적용
            if category in category_colors:
                for cell in row:
                    cell.fill = category_colors[category]
            
            # 우선순위별 글자색 적용
            if priority in priority_colors:
                worksheet[f'G{row_num}'].font = priority_colors[priority]
            
            # 셀 정렬 및 텍스트 래핑
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 열 너비 조정
        column_widths = {
            'A': 8,   # TC ID
            'B': 15,  # Category
            'C': 40,  # Title
            'D': 30,  # Precondition
            'E': 50,  # Test Step
            'F': 40,  # Expected Result
            'G': 10   # Priority
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 행 높이 조정 (자동 조정)
        for row in worksheet.iter_rows(min_row=2):
            max_lines = max(len(str(cell.value).split('\n')) if cell.value else 1 for cell in row)
            worksheet.row_dimensions[row[0].row].height = max(15 * max_lines, 20)
    
    print(f"Excel 파일이 생성되었습니다: {excel_file}")
    return excel_file

if __name__ == "__main__":
    create_test_cases_excel()
